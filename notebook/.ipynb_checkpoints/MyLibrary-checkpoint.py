# ---
# jupyter:
#   jupytext:
#     text_representation:
#       extension: .py
#       format_name: percent
#       format_version: '1.3'
#       jupytext_version: 1.16.2
#   kernelspec:
#     display_name: Python 3 (ipykernel)
#     language: python
#     name: python3
# ---

# %%
# テンプレートExcelの書式に従ってDataFrameのデータを流し込んだExcelを作成する その3
# データの流し込み開始位置指定可。(左上セルを指定)
# Dataframeの列のうちテンプレートの見出しに存在するカラム列のみを流し込むか、DataFrameのカラム名でテンプレートの見出しを上書きして流し込むか、を選択できるようにする。
# 
# df: Dataframe
# format_path: データの見出し(1行目)とデータ書式(2行目)のみインプットされたExcelファイルのパス
# out_path: 結果出力のパス
# start_cell : 流し込み開始位置(見出しの左上のセルの名前)、デフォルトは"A1"
# use_template_columns(True/False) : True-テンプレートの見出しに存在する列のみを流し込む、False-テンプレートの見出しをDataFrameのカラム名で上書きして流し込みする、デフォルトはTrue

def Save_Excel_on_Format(df, format_path, out_path, start_cell="A1", use_template_columns=True):
    import pandas as pd
    import openpyxl as xl
    import re
    from copy import copy
    
    template_wb = xl.load_workbook(filename = format_path)
    template_ws = template_wb.worksheets[0]
    
    x_diff, y_diff = convert_A1_to_xy(start_cell)

    if use_template_columns==True:
        # Trueの場合は、テンプレートの見出しに沿ってデータを配置
        new_df = pd.DataFrame()
        template_columns = get_template_headers(template_ws, start_cell)
        for col in template_columns:
            if col in df.columns:
                new_df[col] = df[col]
            else:
                new_df[col] = ""
    else:
        # Falseの場合は、テンプレートの見出し名を上書き
        new_df = df
        x = x_diff - 1
        y = y_diff
        for col in new_df.columns:
            x = x + 1
            template_ws.cell(row = y, column = x).value = col
        
    for row in range(len(new_df)):
        record = new_df.loc[row]
        x = x_diff - 1
        y = y_diff - 1 + row + 2
        for data in record:
            x = x + 1
            template_ws.cell(row = y, column = x).value = data
            template_ws.cell(row = y, column = x)._style = copy(template_ws.cell(row = y_diff + 1, column = x)._style)
    return template_wb.save(out_path)

# A1スタイルのセル名を(column, row)の数値に変換。例:B3 → (2,3)
def convert_A1_to_xy(cell_name:str):
    import re
    import openpyxl as xl
    
    cell_str = re.search(r"[a-zA-Z]+",cell_name).group()
    cell_num = re.search(r"[0-9]+",cell_name).group()
    if cell_name != cell_str + cell_num:
        return print("start cellの指定が間違っています")
    return xl.utils.column_index_from_string(cell_str), int(cell_num)

# テンプレートExcel内の見出しを取得し、リストで返す
# 見出しに空白や重複があると問題が生じるので注意すること。
def get_template_headers(ws, start_cell="A1"):
    import openpyxl as xl
    start_x, start_y = convert_A1_to_xy(start_cell)  
    for row in ws.iter_rows(min_row=start_y, min_col=start_x, max_row=start_y):
        template_columns = []
        for cell in row:
            # print("({},{})={},書式:{}".format(cell.column,cell.row,cell.value,cell.number_format))
            if cell.value=="":
                print("見出しがない項目があります")
                return []
            template_columns.append(cell.value)
    # # 見出しに重複があるか?
    if len(template_columns) != len(set(template_columns)):
        print("テンプレートの見出しに重複があります")
        return []    
    return template_columns
